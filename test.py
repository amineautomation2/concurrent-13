import time
from re import findall
from camoufox.sync_api import Camoufox
from playwright.sync_api import Page
import openpyxl

def main(url, wb):
    url = url
    wb = wb
    y = 0
    with Camoufox(headless=False) as browser:
        page = browser.new_page()
        page.goto(url)
        scrape(page, url)

def txt(page: Page, xpath: str) -> str:
    try:
        page.wait_for_selector(f"xpath={xpath}", timeout=10000)
        page.wait_for_selector(f"xpath={xpath}")
    except:
        page.wait_for_selector(f"xpath={xpath}")
    
    locate = page.locator(f"xpath={xpath}")
    return locate.text_content() or ""

def element_dropdown(page:Page,xpath: str, option: str) -> None:
    # Locates all matching elements matching the selector
    list_drop_down = page.locator(f"xpath={xpath}").all()
    for drop_down_element in list_drop_down:
        # Playwright handles select options natively via select_option
        drop_down_element.select_option(value=option)
    time.sleep(1)
    
def scroll_down(page: Page, y: int):
    y += 150
    # Evaluates script directly on the active page context
    page.evaluate(f"window.scrollTo(0, {y});")

def click(page:Page, xpath: str) -> None:
    """Helper method preserving original structure click implementation"""

def scrape(page: Page, page_url: str):
    print('[#] Scraping Aviva [#]')
    iter = 2
    wb = openpyxl.load_workbook("")
    ws = wb["aviva"]
    
    # Navigate to target
    page.goto(page_url)

    cookie_xpath = '//*[@id="onetrust-accept-btn-handler"]'
    page.locator(f"xpath={cookie_xpath}").click()

    for x in range(1, 2):
        u = f"https://www.direct.aviva.co.uk/wealth/InvestmentChoice/InvestmentTrustSearch?page={x}"
        page.goto(u)
        print(f'[#] Scraping Page {x} out of 500 [#]')
        
        # Synchronous explicit wait
        page.wait_for_selector('xpath=//*[@id="paginatedResults"]')
        list_funds = []
        
        # Fetch locator array for structural iterations
        table = page.locator('xpath=//*[@id="paginatedResults"]/fieldset/div').all()

        for row in table:
            # Scrape nested text elements relative to parent locator node
            name = row.locator('xpath=./div[2]/label/span/span').text_content()
            url = row.locator('xpath=./div[2]/div[1]/div/a')
            if url: 
                sedol = findall(r'[A-Z0-9]{7}', url.text_content() or "")[0]
                list_funds.append({"name": name, "sedol": sedol})
                print(f'{sedol} | {name}')
#
#        for fund in list_funds:
#            ws.cell(iter, 1).value = fund["sedol"]
#            ws.cell(iter, 2).value = fund["name"]
#            iter += 1
#
#        wb.save('aviva.xlsx')
#        time.sleep(1)
#        
#    wb.save('aviva.xlsx')
#    wb.close()
#       # Cleanly closes the anti-detect browser engine instance
#   
#
url = "https://www.direct.aviva.co.uk/wealth/InvestmentChoice/InvestmentTrustSearch"
main(url, "")